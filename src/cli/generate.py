#!/usr/bin/env python3
import os, sys, json, glob, argparse
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem,
    Table, TableStyle, HRFlowable, KeepTogether
)
from core.config import get_store_file
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from reportlab.pdfbase.pdfmetrics import stringWidth

# =========================
# CONFIG
# =========================
EXCEL_HEADERS = [
    "Dir", "role", "company", "location",
    "jobType", "salary", "link", "address",
    "status", "applyDateTime"
]

# =========================
# UTILITIES
# =========================
def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe_name_from_json(data: dict, fallback_base: str) -> str:
    """Use 'name' from JSON, or fallback to filename, removing spaces."""
    name = (data.get("name") or "").strip()
    if not name:
        name = os.path.splitext(os.path.basename(fallback_base))[0]
    return name.replace(" ", "")

def format_apply_datetime(dt_str: str) -> str:
    """Convert ISO or string datetime to 'Sun Oct 12 2025 19:15' format."""
    if not dt_str:
        return ""
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        return dt.strftime("%a %b %d %Y %H:%M")
    except Exception:
        return str(dt_str)

# =========================
# EXCEL HANDLING
# =========================
def open_or_create_excel(excel_path: str):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(EXCEL_HEADERS)
    return wb, ws

def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = "" if row[0] is None else str(row[0])
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

def append_applydetail_rows(dir_path: str, ws) -> int:
    """Append applyDetail rows from all JSONs in dir_path to Excel."""
    dir_name = os.path.basename(os.path.abspath(dir_path))
    json_files = sorted(glob.glob(os.path.join(dir_path, "*.json")))
    added = 0
    for path in json_files:
        try:
            data = load_json(path)
            ad = data.get("applyDetail") or {}
            if not any(ad.get(k) for k in
                       ["role","company","location","jobType","salary","applyDateTime","status","link","address"]):
                continue

            row = [
                dir_name,
                ad.get("role", ""),
                ad.get("company", ""),
                ad.get("location", ""),
                ad.get("jobType", ""),
                ad.get("salary", ""),
                ad.get("link", ""),
                ad.get("address", ""),
                ad.get("status", ""),
                format_apply_datetime(ad.get("applyDateTime", "")),
            ]
            ws.append(row)
            added += 1
        except Exception as e:
            print(f"⚠️  Skipping '{os.path.basename(path)}' for Excel: {e}", file=sys.stderr)
    return added

# =========================
# PDF GENERATION
# =========================
def section_rule():
    return HRFlowable(width="100%", thickness=0.8, lineCap='round',
                      color=colors.HexColor("#DDDDDD"), spaceBefore=10, spaceAfter=10)

def thin_divider():
    return HRFlowable(width="100%", thickness=0.5, lineCap='round',
                      color=colors.HexColor("#DDDDDD"), spaceBefore=6, spaceAfter=6)

def format_date_range(start, end):
    def fmt(part):
        if isinstance(part, dict):
            y = str(part.get('year', '')).strip()
            m = str(part.get('month', '')).strip()
            return f"{y}/{m}" if y and m else y or m or ""
        return str(part).strip()
    s = fmt(start)
    e = fmt(end)
    return f"{s} - {e}" if s or e else ""

def calculate_duration(start, end):
    start_date = datetime(int(start["year"]), int(start["month"]), 1)
    if isinstance(end, dict) and "present" not in str(end).lower():
        end_date = datetime(int(end["year"]), int(end["month"]), 1)
    else:
        end_date = datetime.today()
    delta_months = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
    years = delta_months // 12
    months = delta_months % 12
    parts = []
    if years:
        parts.append(f"{years} yr{'s' if years > 1 else ''}")
    if months:
        parts.append(f"{months} mo{'s' if months > 1 else ''}")
    return " ".join(parts) if parts else "0 mos"

def is_bullet_line(line: str) -> bool:
    l = line.strip()
    return l.startswith("- ") or l.startswith("• ")

def make_bullet_list(text: str, para_style: ParagraphStyle):
    items = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if is_bullet_line(line):
            line = line[2:].strip()
            items.append(ListItem(Paragraph(line, para_style), leftIndent=6))
    if not items:
        return None
    return ListFlowable(
        items,
        bulletType='bullet',
        bulletFontName='Helvetica',
        bulletFontSize=9,
        bulletColor=colors.HexColor("#333333"),
        leftIndent=10,
        spaceBefore=2,
        spaceAfter=2
    )

def render_detail(text: str, para_style: ParagraphStyle):
    if any(is_bullet_line(l) for l in text.splitlines()):
        lf = make_bullet_list(text, para_style)
        return [lf] if lf else []
    blocks = [b.strip() for b in text.replace("\r\n", "\n").split("\n\n") if b.strip()]
    if not blocks:
        blocks = [b.strip() for b in text.split("\n") if b.strip()]
    return [Paragraph(b, para_style) for b in blocks]

def create_cv_pdf(data: dict, output_pdf_path: str):
    doc = SimpleDocTemplate(output_pdf_path, pagesize=LETTER,
                            rightMargin=50, leftMargin=50,
                            topMargin=60, bottomMargin=40)
    styles = getSampleStyleSheet()
    content = []

    header_style = ParagraphStyle('Header', parent=styles['Heading1'], fontSize=22,
                                  alignment=TA_LEFT, spaceAfter=6)
    role_style = ParagraphStyle('Role', parent=styles['Heading2'], fontSize=12,
                                alignment=TA_LEFT, spaceAfter=12)
    section_header = ParagraphStyle('SectionHeader', parent=styles['Heading2'], fontSize=14,
                                    spaceBefore=6, spaceAfter=6, fontName="Helvetica-Bold")
    normal = styles['Normal']
    normal.leading = normal.fontSize + 2
    exp_text = ParagraphStyle('ExpText', parent=normal, leading=normal.fontSize + 3, spaceBefore=1, spaceAfter=3)
    bullet_text = ParagraphStyle('BulletText', parent=normal, leftIndent=0, leading=normal.fontSize + 2, spaceAfter=1)

    # Header
    content.append(Paragraph(data.get('name', ''), header_style))
    content.append(Paragraph(data.get('role', ''), role_style))

  

    # Contact
    contact = { (k.lower() if isinstance(k, str) else k): v for k, v in (data.get('contact', {}) or {}).items() }
    contact_link_style = ParagraphStyle('ContactLink', parent=normal, alignment=TA_LEFT, fontSize=9)
    contact_text_style = ParagraphStyle('ContactText', parent=normal, alignment=TA_LEFT, fontSize=9)

    line = []
    linkedin = (contact.get("linkedin") or "").strip()
    if linkedin:
        line.append(f"LinkedIn: {linkedin}")
    email = (contact.get("email") or "").strip()
    if email:
        line.append(f"Email: {email}")
    address = (contact.get("address") or "").strip()
    if address:
        line.append(f"Address: {address}")
    phone = (contact.get("phone") or "").strip()
    if phone:
        line.append(f"Mobile: {phone}")

    if line:
        content.append(Paragraph("<br/>".join(line), contact_text_style))


    # Summary
    content.append(section_rule())
    content.append(Paragraph("Summary", section_header))
    content.append(Paragraph(data.get("summary", "") or "", normal))

    # Core Skills
    content.append(section_rule())
    content.append(Paragraph("Skills", section_header))
    for group in data.get("coreSkills", []) or []:
        cat = group.get("category", "")
        skills_line = ", ".join(group.get("skills", []) or [])
        content.append(Paragraph(f"<b>{cat}:</b> {skills_line}", normal))
    content.append(Spacer(1, 0.12 * inch))

    # Experience
    content.append(section_rule())
    content.append(Paragraph("Experience", section_header))
    for i, exp in enumerate(data.get("experiences", []) or []):
        date_range = format_date_range(exp.get('start', {}), exp.get('end', {}))
        duration = calculate_duration(exp.get('start', {}), exp.get('end', {}))
        right_text = f"{date_range} ({duration})"
        right_w = stringWidth(right_text, normal.fontName, normal.fontSize) + 6
        t = Table(
            [[Paragraph(f"<b>{exp.get('role','')}</b>", normal),
              Paragraph(f"<b>{right_text}</b>", normal)]],
            colWidths=["*", right_w], hAlign="LEFT"
        )
        t.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        company_line = f"<b>{exp.get('company','')}</b>, {exp.get('location','')} . {exp.get('type','')} . {exp.get('workType','')}"
        detail_flow = render_detail(exp.get('detail', '') or '', bullet_text)
        content.append(KeepTogether([t, Paragraph(company_line, exp_text)] + detail_flow))
        if i < len((data.get("experiences") or [])) - 1:
            content.append(thin_divider())

    # Education
    content.append(section_rule())
    content.append(Paragraph("Education", section_header))
    edu = data.get("education", {}) or {}
    edu_left = f"<b>{edu.get('grade','')}</b>"
    edu_range = ""
    if edu.get("start") or edu.get("end"):
        edu_range = format_date_range(edu.get('start'), edu.get('end'))
    if edu_range:
        right_text = f"{edu_range}"
        right_w = stringWidth(right_text, normal.fontName, normal.fontSize) + 6
        edu_table = Table(
            [[Paragraph(edu_left, normal), Paragraph(f"<b>{right_text}</b>", normal)]],
            colWidths=["*", right_w], hAlign="LEFT"
        )
        edu_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        content.append(edu_table)
    else:
        content.append(Paragraph(edu_left, normal))
    content.append(Paragraph(edu.get('university', ''), normal))

    # Languages
    content.append(section_rule())
    content.append(Paragraph("Languages", section_header))
    for lang in data.get("languages", []) or []:
        content.append(Paragraph(f"{lang.get('language','')}: {lang.get('level','')}", normal))

    content.append(section_rule())
    doc.build(content)
    print(f"🧾 PDF written: {output_pdf_path}")

# =========================
# MAIN
# =========================
def main():
    parser = argparse.ArgumentParser(
        description="Generate PDFs and append applyDetail data into a global Excel file."
    )
    parser.add_argument("--dir", "-d", required=True, help="Directory containing JSON files")
    parser.add_argument("--excel-name", default="All_applyDetail.xlsx", help="Excel file name stored next to this script")
    args = parser.parse_args()

    # in_dir = args.dir
    in_dir = os.path.abspath(os.path.join(project_root, in_dir))
    if not os.path.isdir(in_dir):
        print(f"❌ '{in_dir}' is not a directory.", file=sys.stderr)
        sys.exit(1)

    json_files = sorted(glob.glob(os.path.join(in_dir, "*.json")))
    if not json_files:
        print(f"⚠️ No .json files found in '{in_dir}'.")
        sys.exit(0)

    # Excel setup
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # excel_path = os.path.join(script_dir, args.excel_name)
    # Save Excel in the data/ directory at project root
    project_root = os.path.abspath(os.path.join(script_dir, ".."))
    # excel_path = os.path.join(project_root, "data", args.excel_name)
    excel_path = str(get_store_file())
    wb, ws = open_or_create_excel(excel_path)

    # Generate PDFs
    pdf_count = 0
    for path in json_files:
        try:
            data = load_json(path)
            pdf_name = f"{safe_name_from_json(data, path)}.pdf"
            out_pdf = os.path.join(os.path.dirname(path), pdf_name)
            create_cv_pdf(data, out_pdf)
            pdf_count += 1
        except Exception as e:
            print(f"⚠️ Skipping PDF for '{os.path.basename(path)}': {e}", file=sys.stderr)

    # Append Excel rows
    added = append_applydetail_rows(in_dir, ws)
    if added:
        autosize_columns(ws)
        wb.save(excel_path)
        print(f"✅ Added {added} rows → {excel_path}")
    else:
        wb.save(excel_path)
        print(f"⚠️ No applyDetail found; Excel ensured at {excel_path}")

    print(f"✅ PDFs created: {pdf_count}")

if __name__ == "__main__":
    main()
